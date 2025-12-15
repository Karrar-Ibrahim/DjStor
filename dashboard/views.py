from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import permission_required, user_passes_test
from django.contrib import messages
from django.db import models
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncMonth
from django.utils import timezone
from django.http import HttpResponse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# استيراد المودلز
from store.models import Product, Category, Order, Coupon
from django.contrib.auth.models import User
from store.models import ProductImage # تأكد من وجود هذا الاستيراد
# استيراد الفورم
from .forms import ProductForm, CategoryForm, CouponForm, StaffUserForm

from store.models import HomeSection
from .forms import HomeSectionForm









# --- الصفحة الرئيسية للوحة التحكم ---
@staff_member_required
def dashboard_home(request):
    today = timezone.now().date()
    total_products = Product.objects.count()
    orders_today = Order.objects.filter(created_at__date=today).count()
    total_sales = Order.objects.filter(status='completed').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_orders = Order.objects.count()

    context = {
        'total_products': total_products,
        'orders_today': orders_today,
        'total_sales': total_sales,
        'total_orders': total_orders,
    }
    return render(request, 'dashboard/index.html', context)

# --- إدارة المستخدمين والصلاحيات (للمدير العام فقط) ---
@user_passes_test(lambda u: u.is_superuser)
def users_list(request):
    users = User.objects.filter(is_staff=True).exclude(id=request.user.id)
    return render(request, 'dashboard/users.html', {'users': users})

@user_passes_test(lambda u: u.is_superuser)
def user_add(request):
    if request.method == 'POST':
        form = StaffUserForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "تم إضافة الموظف وتحديد صلاحياته.")
            return redirect('dashboard_users')
    else:
        form = StaffUserForm()
    return render(request, 'dashboard/user_form.html', {'form': form, 'title': 'إضافة موظف جديد'})

@user_passes_test(lambda u: u.is_superuser)
def user_edit(request, pk):
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = StaffUserForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, "تم تحديث بيانات وصلاحيات الموظف.")
            return redirect('dashboard_users')
    else:
        form = StaffUserForm(instance=user)
    return render(request, 'dashboard/user_form.html', {'form': form, 'title': f'تعديل صلاحيات: {user.username}'})

@user_passes_test(lambda u: u.is_superuser)
def user_delete(request, pk):
    user = get_object_or_404(User, pk=pk)
    user.delete()
    messages.success(request, "تم حذف الحساب.")
    return redirect('dashboard_users')

# --- إدارة المنتجات ---
@staff_member_required
def product_manage(request):
    products = Product.objects.all()
    return render(request, 'dashboard/products.html', {'products': products})

@permission_required('store.add_product', raise_exception=True)
def product_add(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save()
            # كود حفظ الصور الإضافية إن وجد
            if 'more_images' in request.FILES:
                from store.models import ProductImage
                images = request.FILES.getlist('more_images')
                for img in images:
                    ProductImage.objects.create(product=product, image=img)
            
            messages.success(request, "تم إضافة المنتج بنجاح.")
            return redirect('dashboard_products')
    else:
        form = ProductForm()
    return render(request, 'dashboard/product_form.html', {'form': form, 'title': 'إضافة منتج جديد'})

@permission_required('store.change_product', raise_exception=True)
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            # كود حفظ الصور الإضافية
            if 'more_images' in request.FILES:
                from store.models import ProductImage
                images = request.FILES.getlist('more_images')
                for img in images:
                    ProductImage.objects.create(product=product, image=img)

            messages.success(request, "تم تعديل المنتج بنجاح.")
            return redirect('dashboard_products')
    else:
        form = ProductForm(instance=product)
    
    return render(request, 'dashboard/product_form.html', {
        'form': form, 
        'title': f'تعديل المنتج: {product.name}'
    })

@permission_required('store.delete_product', raise_exception=True)
def delete_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    product.delete()
    messages.success(request, "تم حذف المنتج بنجاح.")
    return redirect('dashboard_products')

# --- إدارة الفئات (الأقسام) ---
@staff_member_required
def category_list(request):
    categories = Category.objects.all()
    return render(request, 'dashboard/categories.html', {'categories': categories})

@permission_required('store.add_category', raise_exception=True)
def category_add(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "تم إضافة القسم بنجاح.")
            return redirect('dashboard_categories')
    else:
        form = CategoryForm()
    return render(request, 'dashboard/category_form.html', {'form': form, 'title': 'إضافة قسم جديد'})

@permission_required('store.change_category', raise_exception=True)
def category_edit(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, "تم تعديل القسم بنجاح.")
            return redirect('dashboard_categories')
    else:
        form = CategoryForm(instance=category)
    return render(request, 'dashboard/category_form.html', {'form': form, 'title': f'تعديل القسم: {category.name}'})

@permission_required('store.delete_category', raise_exception=True)
def delete_category(request, pk):
    category = get_object_or_404(Category, pk=pk)
    category.delete()
    messages.success(request, "تم حذف الفئة بنجاح.")
    return redirect('dashboard_categories')

# --- إدارة الكوبونات ---
@staff_member_required
def coupon_list(request):
    coupons = Coupon.objects.all().order_by('-valid_to')
    return render(request, 'dashboard/coupons.html', {'coupons': coupons})

@permission_required('store.add_coupon', raise_exception=True)
def coupon_add(request):
    if request.method == 'POST':
        form = CouponForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "تم إضافة الكوبون بنجاح.")
            return redirect('dashboard_coupons')
    else:
        form = CouponForm()
    return render(request, 'dashboard/coupon_form.html', {'form': form, 'title': 'إضافة كود خصم جديد'})

@permission_required('store.change_coupon', raise_exception=True)
def coupon_edit(request, pk):
    coupon = get_object_or_404(Coupon, pk=pk)
    if request.method == 'POST':
        form = CouponForm(request.POST, instance=coupon)
        if form.is_valid():
            form.save()
            messages.success(request, "تم تعديل بيانات الكوبون بنجاح.")
            return redirect('dashboard_coupons')
    else:
        form = CouponForm(instance=coupon)
    return render(request, 'dashboard/coupon_form.html', {'form': form, 'title': f'تعديل الكوبون: {coupon.code}'})

@permission_required('store.delete_coupon', raise_exception=True)
def coupon_delete(request, pk):
    coupon = get_object_or_404(Coupon, pk=pk)
    coupon.delete()
    messages.success(request, "تم حذف الكوبون.")
    return redirect('dashboard_coupons')

# --- إدارة الطلبات ---
@staff_member_required
def order_manage(request):
    orders = Order.objects.all().order_by('-created_at')
    return render(request, 'dashboard/orders.html', {'orders': orders})

@staff_member_required
def order_detail(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    if request.method == 'POST':
        # التحقق من صلاحية تغيير الطلب قبل التنفيذ
        # يمكنك تفعيل هذا السطر إذا أردت تقييد تعديل الطلبات أيضاً
        # if not request.user.has_perm('store.change_order'):
        #      messages.error(request, "ليس لديك صلاحية لتعديل حالة الطلب.")
        #      return redirect('dashboard_order_detail', order_id=order.id)

        status = request.POST.get('status')
        if status in ['pending', 'completed', 'cancelled']:
            order.status = status
            order.save()
            messages.success(request, "تم تحديث حالة الطلب.")
            return redirect('dashboard_orders')
    return render(request, 'dashboard/order_detail.html', {'order': order})

# --- التقارير والجرد ---
@staff_member_required
def dashboard_reports(request):
    selected_year = request.GET.get('year')
    selected_month = request.GET.get('month')
    
    orders_query = Order.objects.filter(status='completed')

    if selected_year and selected_year != 'all':
        orders_query = orders_query.filter(created_at__year=selected_year)
    
    if selected_month and selected_month != 'all':
        orders_query = orders_query.filter(created_at__month=selected_month)

    totals = orders_query.aggregate(
        total_revenue=Sum('total_amount'),
        total_count=Count('id')
    )

    if selected_month and selected_month != 'all':
        sales_data = orders_query.extra(select={'day': "date(created_at)"}).values('day')\
            .annotate(revenue=Sum('total_amount'), count=Count('id')).order_by('-day')
        report_type = 'daily'
    else:
        sales_data = orders_query.annotate(period=TruncMonth('created_at'))\
            .values('period')\
            .annotate(revenue=Sum('total_amount'), count=Count('id')).order_by('-period')
        report_type = 'monthly'

    products_inventory = Product.objects.all().annotate(
        total_sold=Sum('orderitem__quantity', filter=models.Q(orderitem__order__status='completed'))
    ).order_by('stock_quantity')

    available_years = Order.objects.dates('created_at', 'year')

    context = {
        'sales_data': sales_data,
        'products_inventory': products_inventory,
        'totals': totals,
        'selected_year': selected_year,
        'selected_month': selected_month,
        'available_years': available_years,
        'report_type': report_type,
        'current_date': timezone.now()
    }
    return render(request, 'dashboard/reports.html', context)

@staff_member_required
def dashboard_inventory(request):
    products = Product.objects.all().annotate(
        total_sold=Sum('orderitem__quantity', filter=models.Q(orderitem__order__status='completed'))
    ).order_by('stock_quantity')

    context = {
        'products': products,
        'current_date': timezone.now(),
        'total_products_count': products.count(),
        'total_stock_value': sum((p.stock_quantity * p.price) for p in products)
    }
    return render(request, 'dashboard/inventory.html', context)

@staff_member_required
def export_reports_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Reports-{datetime.now().strftime("%Y-%m-%d")}.xlsx'

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "تقرير المبيعات"
    ws1.sheet_view.rightToLeft = True

    headers = ['رقم الطلب', 'العميل', 'التاريخ', 'الإجمالي', 'مبلغ التوصيل', 'الخصم']
    ws1.append(headers)
    
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    orders = Order.objects.filter(status='completed').order_by('-created_at')
    for order in orders:
        ws1.append([
            order.id,
            order.full_name,
            order.created_at.strftime("%Y-%m-%d %H:%M"),
            order.total_amount,
            order.delivery_fee,
            order.discount_amount
        ])

    ws2 = wb.create_sheet(title="جرد المخزون")
    ws2.sheet_view.rightToLeft = True

    headers_inv = ['المنتج', 'القسم', 'السعر', 'الكمية المتبقية', 'الحالة']
    ws2.append(headers_inv)

    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    products = Product.objects.all().order_by('stock_quantity')
    for p in products:
        status = "نفذت الكمية" if p.stock_quantity == 0 else "متوفر"
        ws2.append([
            p.name,
            p.category.name,
            p.final_price,
            p.stock_quantity,
            status
        ])

    wb.save(response)
    return response




@permission_required('store.change_product', raise_exception=True)
def delete_product_image(request, image_id):
    # جلب الصورة
    img = get_object_or_404(ProductImage, id=image_id)
    product_id = img.product.id # نحتفظ برقم المنتج لنعود إليه
    
    # حذف الصورة من الملفات ومن قاعدة البيانات
    img.delete()
    
    messages.success(request, "تم حذف الصورة من المعرض.")
    return redirect('dashboard_product_edit', pk=product_id)




@permission_required('store.change_product', raise_exception=True)
def delete_main_image(request, pk):
    product = get_object_or_404(Product, pk=pk)
    
    # حذف الصورة من الملفات إذا كانت موجودة
    if product.main_image:
        product.main_image.delete(save=False)
        product.main_image = None # تفريغ الحقل في قاعدة البيانات
        product.save()
        messages.success(request, "تم حذف الصورة الرئيسية للمنتج.")
    
    return redirect('dashboard_product_edit', pk=pk)



@staff_member_required
def home_sections_list(request):
    sections = HomeSection.objects.all()
    return render(request, 'dashboard/home_sections.html', {'sections': sections})

@staff_member_required
def home_section_add(request):
    if request.method == 'POST':
        form = HomeSectionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "تم إضافة السكشن بنجاح.")
            return redirect('dashboard_home_sections')
    else:
        form = HomeSectionForm()
    return render(request, 'dashboard/home_section_form.html', {'form': form})

@staff_member_required
def home_section_delete(request, pk):
    section = get_object_or_404(HomeSection, pk=pk)
    section.delete()
    messages.success(request, "تم حذف السكشن.")
    return redirect('dashboard_home_sections')



@staff_member_required
def home_section_edit(request, pk):
    section = get_object_or_404(HomeSection, pk=pk)
    if request.method == 'POST':
        form = HomeSectionForm(request.POST, instance=section)
        if form.is_valid():
            form.save()
            messages.success(request, "تم تحديث السكشن بنجاح.")
            return redirect('dashboard_home_sections')
    else:
        form = HomeSectionForm(instance=section)
    
    return render(request, 'dashboard/home_section_form.html', {
        'form': form,
        'title': f'تعديل السكشن: {section.title}'
    })